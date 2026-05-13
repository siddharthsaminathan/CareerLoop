"""
CareerLoop Excel Audit — Generate audit reports.

Not the primary UX. The audit layer for review and accountability.
"""

import csv
import os
from datetime import datetime, timezone
from pathlib import Path


class AuditReport:
    """Generate Excel/CSV audit reports for job pipeline."""

    COLUMNS = [
        "person",
        "source",
        "company",
        "role",
        "url",
        "location",
        "work_mode",
        "salary",
        "company_type",
        "fit_score",
        "recommendation",
        "why_fit",
        "risks",
        "status",
        "decision",
        "skip_reason",
        "first_seen_at",
        "last_seen_at",
        "follow_up_at",
        "next_action",
    ]

    def __init__(self, career_ops_root: str):
        self.root = Path(career_ops_root)
        self.reports_dir = self.root / "reports"

    def generate(self, ledger_entries: list[dict], person_name: str = "",
                 date: str = None) -> str:
        """Generate audit CSV. Returns file path."""
        if date is None:
            date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        report_dir = self.reports_dir / date
        report_dir.mkdir(parents=True, exist_ok=True)

        safe_name = person_name.lower().replace(" ", "-") if person_name else "user"
        filename = f"{safe_name}-job-report.csv"
        filepath = report_dir / filename

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(self.COLUMNS)

            for entry in ledger_entries:
                fit = entry.get("fit_result", {})
                row = [
                    person_name,
                    entry.get("source", ""),
                    entry.get("company", ""),
                    entry.get("title", entry.get("role_title", "")),
                    entry.get("source_url", entry.get("application_url", "")),
                    entry.get("location", ""),
                    entry.get("work_mode", ""),
                    entry.get("salary_range", ""),
                    entry.get("company_type", ""),
                    entry.get("fit_score") or fit.get("overall_score", ""),
                    fit.get("recommendation", ""),
                    fit.get("why_user_might_like_it", fit.get("reason", "")),
                    " | ".join(fit.get("risks", [])),
                    entry.get("status", ""),
                    entry.get("user_decision", entry.get("decision", "")),
                    entry.get("skip_reason", ""),
                    entry.get("first_seen_at", entry.get("created_at", "")),
                    entry.get("last_seen_at", entry.get("updated_at", "")),
                    entry.get("follow_up_dates", [None])[0] if entry.get("follow_up_dates") else "",
                    entry.get("next_action", ""),
                ]
                writer.writerow(row)

        return str(filepath)

    def generate_xlsx(self, ledger_entries: list[dict], person_name: str = "",
                      date: str = None) -> str:
        """Generate Excel (.xlsx) audit report. Returns file path."""
        try:
            import openpyxl
        except ImportError:
            return self.generate(ledger_entries, person_name, date)  # fallback to CSV

        if date is None:
            date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        report_dir = self.reports_dir / date
        report_dir.mkdir(parents=True, exist_ok=True)

        safe_name = person_name.lower().replace(" ", "-") if person_name else "user"
        filename = f"{safe_name}-job-report.xlsx"
        filepath = report_dir / filename

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Pipeline"

        # Header
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data
        for row_idx, entry in enumerate(ledger_entries, 2):
            fit = entry.get("fit_result", {})
            values = [
                person_name,
                entry.get("source", ""),
                entry.get("company", ""),
                entry.get("title", entry.get("role_title", "")),
                entry.get("source_url", ""),
                entry.get("location", ""),
                entry.get("work_mode", ""),
                entry.get("salary_range", ""),
                entry.get("company_type", ""),
                entry.get("fit_score") or fit.get("overall_score", ""),
                fit.get("recommendation", ""),
                fit.get("why_user_might_like_it", ""),
                " | ".join(fit.get("risks", [])),
                entry.get("status", ""),
                entry.get("user_decision", ""),
                entry.get("skip_reason", ""),
                entry.get("first_seen_at", entry.get("created_at", "")),
                entry.get("last_seen_at", entry.get("updated_at", "")),
                str(entry.get("follow_up_dates", [None])[0]) if entry.get("follow_up_dates") else "",
                entry.get("next_action", ""),
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=val)

        # Auto-width
        for col in range(1, len(self.COLUMNS) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        wb.save(filepath)
        return str(filepath)
